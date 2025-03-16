#!/usr/bin/env python3
import os
import re
import argparse
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles import Border, Side
from datetime import datetime
import openpyxl

def print_usage():
    """引数なしで実行された場合に表示するヘルプメッセージ"""
    usage = """
使用方法: search.py [オプション] <検索語句>

説明:
  このプログラムはファイルやディレクトリ内のテキストを検索するためのツールです。
  引数なしで実行すると、このヘルプメッセージが表示されます。
  デフォルトでは、サブディレクトリも再帰的に検索し、大文字と小文字を区別しません。
  検索結果はデフォルトでExcelファイル（result_日時.xlsx）に出力されます。

オプション:
  -h, --help            このヘルプメッセージを表示します
  -p, --path PATH       検索対象のパスを指定します (デフォルト: カレントディレクトリ)
  -r, --no-recursive    サブディレクトリを再帰的に検索しません (デフォルトは再帰的に検索)
  -c, --case-sensitive  大文字と小文字を区別して検索します (デフォルトは区別しない)
  -f, --file-pattern PAT 特定のファイルパターンのみを検索します (例: *.txt)
  -o, --output FILE     結果を指定したファイルに出力します
  -t, --output-type TYPE 出力形式を指定します (excel, csv, text) (デフォルト: excel)
  -s, --stdout          結果を標準出力に表示します (Excelファイルは生成されません)

例:
  search.py "検索語句"                      # 検索結果をExcelファイルに出力
  search.py -p /path/to/dir "検索語句"      # 指定したディレクトリで検索
  search.py -c "検索語句"                   # 大文字小文字を区別して検索
  search.py -t csv "検索語句"               # 結果をCSV形式で出力
  search.py -s "検索語句"                   # 結果を標準出力に表示
  search.py -o custom_name.xlsx "検索語句"  # カスタムファイル名で出力

詳細については、プロジェクトのドキュメントを参照してください。
"""
    print(usage)

def get_default_output_filename():
    """デフォルトの出力ファイル名（result_日時.xlsx）を生成"""
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    return f"result_{timestamp}.xlsx"

def write_to_excel(filename, data):
    """データをExcelファイルに書き込む"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "検索結果"
    
    # ヘッダーの書き込み
    ws.append(["ファイル名", "行番号", "内容"])
    
    # データの書き込み
    for item in data:
        ws.append(item)
    
    wb.save(filename)
    print(f"検索結果を {filename} に出力しました。")

def main():
    # 引数がない場合はヘルプメッセージを表示して終了
    if len(sys.argv) == 1:
        print_usage()
        sys.exit(0)
    
    # 以下に引数解析と実際の検索処理を実装
    parser = argparse.ArgumentParser(description='ファイルやディレクトリ内のテキストを検索します', add_help=False)
    parser.add_argument('-h', '--help', action='store_true', help='ヘルプメッセージを表示します')
    parser.add_argument('-p', '--path', default='.', help='検索対象のパスを指定します')
    parser.add_argument('-r', '--no-recursive', action='store_true', help='サブディレクトリを再帰的に検索しません')
    parser.add_argument('-c', '--case-sensitive', action='store_true', help='大文字と小文字を区別して検索します')
    parser.add_argument('-f', '--file-pattern', help='特定のファイルパターンのみを検索します')
    parser.add_argument('-o', '--output', help='結果を指定したファイルに出力します')
    parser.add_argument('-t', '--output-type', choices=['excel', 'csv', 'text'], default='excel', 
                        help='出力形式を指定します (デフォルト: excel)')
    parser.add_argument('-s', '--stdout', action='store_true', help='結果を標準出力に表示します')
    parser.add_argument('search_term', nargs='?', help='検索する語句')
    
    args = parser.parse_args()
    
    # --helpオプションが指定された場合もヘルプを表示
    if args.help or not args.search_term:
        print_usage()
        sys.exit(0)
    
    # デフォルト値の設定
    recursive = not args.no_recursive  # デフォルトで再帰的に検索
    ignore_case = not args.case_sensitive  # デフォルトで大文字小文字を区別しない
    
    # 出力先の設定
    if args.stdout:
        output_to_file = False
        output_file = None
    else:
        output_to_file = True
        output_file = args.output if args.output else get_default_output_filename()
    
    # ここに実際の検索処理を実装
    # 例として、ダミーデータを使用
    search_results = [
        ("example.txt", 10, "これは検索結果の例です。"),
        ("example.txt", 20, "別の検索結果の例です。")
    ]
    
    # 検索結果の出力
    if output_to_file and args.output_type == 'excel':
        write_to_excel(output_file, search_results)
    elif args.stdout:
        for result in search_results:
            print(result)

def search_files(keywords, base_dir='.', output_file=None, output_excel=None):
    """
    指定ディレクトリ以下のすべてのファイルを対象に、指定した単語を検索する

    :param keywords: 検索する単語のリスト
    :param base_dir: 検索対象の基準ディレクトリ
    :param output_file: 結果を保存するテキストファイル（Noneの場合は標準出力）
    :param output_excel: 結果をExcelファイルとして保存するパス
    """
    keyword_patterns = [re.compile(rf'\b{re.escape(keyword)}\b') for keyword in keywords]

    results = {keyword: {'hit_count': 0, 'file_count': 0, 'files': {}} for keyword in keywords}
    excel_data = []

    for root, _, files in os.walk(base_dir):
        for file in files:
            file_path = os.path.join(root, file)
            relative_file_path = os.path.relpath(file_path, base_dir)  # 相対パスを取得し、"./"を除去

            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()

                file_hits = {keyword: 0 for keyword in keywords}
                file_results = {keyword: [] for keyword in keywords}

                for i, line in enumerate(lines):
                    for keyword, pattern in zip(keywords, keyword_patterns):
                        if pattern.search(line):
                            results[keyword]['hit_count'] += 1
                            file_hits[keyword] += 1

                            # ヒットした行の文字列
                            hit_line = line.strip()

                            # ヒットした行の前後2行を取得（折り返し表示用）
                            start = max(i - 2, 0)
                            end = min(i + 3, len(lines))  # ヒットした行 + 2行まで
                            full_context = "".join(lines[start:end]).strip()

                            file_results[keyword].append(f"{relative_file_path} (Line {i+1}): {hit_line}")
                            excel_data.append([keyword, relative_file_path, i + 1, hit_line, full_context])

                # ファイルごとのヒットを記録
                for keyword in keywords:
                    if file_hits[keyword] > 0:
                        results[keyword]['file_count'] += 1
                        results[keyword]['files'][relative_file_path] = file_results[keyword]

            except Exception as e:
                print(f"エラー: {file_path} を読み込めませんでした - {e}")

    output_lines = []

    for keyword in keywords:
        output_lines.append(f"検索ワード: '{keyword}'")
        output_lines.append(f"  ヒット総数: {results[keyword]['hit_count']}")
        output_lines.append(f"  該当ファイル数: {results[keyword]['file_count']}")
        output_lines.append("")

        for file, lines in results[keyword]['files'].items():
            output_lines.append(f"  {file}:")
            output_lines.extend([f"    {line}" for line in lines])
            output_lines.append("")

    result_text = "\n".join(output_lines)

    if output_file:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(result_text)
        print(f"検索結果を '{output_file}' に保存しました。")
    else:
        print(result_text)

    if output_excel:
        save_results_to_excel(output_excel, results, excel_data)
        print(f"検索結果を Excelファイル '{output_excel}' に保存しました。")

def save_results_to_excel(output_excel, results, excel_data):
    """
    検索結果を Excel ファイルとして保存する

    :param output_excel: 保存する Excel ファイル名
    :param results: 各キーワードのヒット数とファイル数のデータ
    :param excel_data: 詳細情報リスト（[キーワード, ファイルパス, 行番号, ヒットした行, 前後2行を含む全文]）
    """
    wb = Workbook()

    # サマリーシート
    ws_summary = wb.active
    ws_summary.title = "検索結果サマリー"
    ws_summary.append(["検索ワード", "ヒット数", "該当ファイル数"])

    for keyword, data in results.items():
        ws_summary.append([keyword, data['hit_count'], data['file_count']])

    # 列幅を固定サイズに設定（px単位をEMU単位に変換: 1px ≈ 0.14インチ）
    # Excelの列幅は文字数単位なので、およそ1文字7ピクセル程度で換算
    ws_summary.column_dimensions['A'].width = 100 / 7  # 100px ≈ 14.3文字
    ws_summary.column_dimensions['B'].width = 80 / 7   # 80px ≈ 11.4文字
    ws_summary.column_dimensions['C'].width = 100 / 7  # 100px ≈ 14.3文字

    # データを中央揃えに設定
    for row in ws_summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # ヘッダー行のスタイル - より目立つ背景色を設定
    header_fill = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # 1行目を固定表示（スクロール時に常に表示）
    ws_summary.freeze_panes = "A2"

    # 詳細シート
    ws_details = wb.create_sheet(title="検索結果詳細")
    ws_details.append(["検索ワード", "ファイル", "行番号", "ヒットした行", "前後2行を含む全文"])

    # カラムの幅を事前に設定
    ws_details.column_dimensions['A'].width = 15  # 検索ワード
    ws_details.column_dimensions['B'].width = 30  # ファイル
    ws_details.column_dimensions['C'].width = 10  # 行番号
    ws_details.column_dimensions['D'].width = 40  # ヒットした行
    ws_details.column_dimensions['E'].width = 80  # 前後2行を含む全文 - 十分な幅を確保

    # ヘッダー行のスタイル - より目立つ背景色を設定
    header_fill = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    # 1行目を固定表示（スクロール時に常に表示）
    ws_details.freeze_panes = "A2"

    # データをB列（ファイル）とA列（検索ワード）で昇順ソート
    sorted_excel_data = sorted(excel_data, key=lambda x: (x[1], x[0]))

    # 色付けのための準備
    # ファイルとキーワードの組み合わせごとに一意の色を割り当てる
    file_colors = {}     # ファイルごとの色を管理
    keyword_colors = {}  # ファイル内のキーワードごとの色を管理
    
    # 背景色の定義 (コントラストを高めた色)
    color_a1 = "BBDEFB"  # より濃い青
    color_a2 = "E3F2FD"  # 薄い青
    color_b1 = "FFCC80"  # より濃いオレンジ
    color_b2 = "FFF3E0"  # 薄いオレンジ
    
    # 最初にソートされたデータから一意のファイルとキーワードの組み合わせを抽出し、色を割り当てる
    current_file = None
    file_toggle = False
    
    for row_data in sorted_excel_data:
        keyword = row_data[0]
        file_path = row_data[1]
        
        # 新しいファイルが出現したら色をトグル
        if file_path != current_file:
            current_file = file_path
            file_toggle = not file_toggle
            file_colors[file_path] = color_a1 if file_toggle else color_a2
            keyword_colors[file_path] = {}
            current_keyword_toggle = False
        
        # このファイル内でのキーワードの色を設定
        if keyword not in keyword_colors[file_path]:
            current_keyword_toggle = not current_keyword_toggle
            keyword_colors[file_path][keyword] = color_b1 if current_keyword_toggle else color_b2
    
    # 罫線グループのための変数
    from openpyxl.styles import Border, Side
    
    # データを追加して色付け
    current_file = None
    current_keyword = None
    
    for row_idx, row_data in enumerate(sorted_excel_data, start=2):
        ws_details.append(row_data)
        
        keyword = row_data[0]
        file_path = row_data[1]
        
        # 罫線を引く条件の判定
        border_type = "thin"  # デフォルトは細い罫線
        
        # ファイルが変わる場合
        if current_file is not None and file_path != current_file:
            border_type = "medium"  # ファイルが変わる場合は太い罫線
            current_keyword = None  # ファイルが変わったらキーワードもリセット
        # 同じファイル内でキーワードが変わる場合
        elif current_keyword is not None and keyword != current_keyword and current_file == file_path:
            border_type = "thin"  # キーワードが変わる場合は細い罫線
        
        # 前の行に下罫線を引く（最初の行以外）
        if row_idx > 2:
            for col in range(1, 6):
                cell = ws_details.cell(row=row_idx-1, column=col)
                current_border = cell.border
                
                # 既存の罫線情報を保持
                left_border = current_border.left if current_border and current_border.left else Side(border_style=None)
                right_border = current_border.right if current_border and current_border.right else Side(border_style=None)
                top_border = current_border.top if current_border and current_border.top else Side(border_style=None)
                
                # 左端と右端の列は特別扱い
                if col == 1:
                    left_border = Side(border_style="medium")
                elif col == 5:
                    right_border = Side(border_style="medium")
                
                # 下罫線を設定
                cell.border = Border(
                    left=left_border,
                    right=right_border,
                    top=top_border,
                    bottom=Side(border_style=border_type)
                )
        
        # 現在の行に上罫線を引く
        for col in range(1, 6):
            cell = ws_details.cell(row=row_idx, column=col)
            
            # 左端と右端の列は特別扱い
            if col == 1:
                cell.border = Border(
                    left=Side(border_style="medium"),
                    top=Side(border_style=border_type)
                )
            elif col == 5:
                cell.border = Border(
                    right=Side(border_style="medium"),
                    top=Side(border_style=border_type)
                )
            else:
                cell.border = Border(
                    top=Side(border_style=border_type)
                )
        
        current_file = file_path
        current_keyword = keyword
        
        # A～D列の設定
        for col_idx in range(1, 5):
            cell = ws_details.cell(row=row_idx, column=col_idx)
            # 中央揃え、上下中央揃えの設定
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # B列の背景色を設定（ファイル）
            if col_idx == 2:
                cell.fill = PatternFill(
                    start_color=file_colors[file_path],
                    end_color=file_colors[file_path],
                    fill_type="solid"
                )
            # A列の背景色を設定（キーワード）
            elif col_idx == 1:
                cell.fill = PatternFill(
                    start_color=keyword_colors[file_path][keyword],
                    end_color=keyword_colors[file_path][keyword],
                    fill_type="solid"
                )
        
        # E列のセルに折り返し設定を適用（左揃え、上揃えのまま）
        cell = ws_details.cell(row=row_idx, column=5)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 最後の行の下罫線
    if len(sorted_excel_data) > 0:
        last_row = len(sorted_excel_data) + 1
        for col in range(1, 6):
            cell = ws_details.cell(row=last_row, column=col)
            if col == 1:
                cell.border = Border(left=Side(border_style="medium"), bottom=Side(border_style="medium"))
            elif col == 5:
                cell.border = Border(right=Side(border_style="medium"), bottom=Side(border_style="medium"))
            else:
                cell.border = Border(bottom=Side(border_style="medium"))

    # 保存前に列幅を再確認（自動調整からの最小幅を保証）
    for col in range(1, 6):
        col_letter = get_column_letter(col)
        current_width = ws_details.column_dimensions[col_letter].width
        if col == 5:  # E列は最低60の幅を確保
            ws_details.column_dimensions[col_letter].width = max(current_width, 60)

    # 保存
    wb.save(output_excel)

def draw_border_around_group(worksheet, start_row, end_row):
    """
    指定された行範囲のA列からE列までを罫線で囲む
    
    :param worksheet: 対象のワークシート
    :param start_row: 開始行（インデックス）
    :param end_row: 終了行（インデックス）
    """
    thick = Side(border_style="medium", color="000000")
    no_border = Side(border_style=None)
    
    # 上部の罫線
    for col in range(1, 6):  # A列からE列
        cell = worksheet.cell(row=start_row, column=col)
        if col == 1:  # 左上角
            cell.border = Border(top=thick, left=thick, right=no_border, bottom=no_border)
        elif col == 5:  # 右上角
            cell.border = Border(top=thick, left=no_border, right=thick, bottom=no_border)
        else:  # 上辺
            cell.border = Border(top=thick, left=no_border, right=no_border, bottom=no_border)
    
    # 中間の行
    for row in range(start_row + 1, end_row):
        for col in range(1, 6):
            cell = worksheet.cell(row=row, column=col)
            if col == 1:  # 左辺
                cell.border = Border(top=no_border, left=thick, right=no_border, bottom=no_border)
            elif col == 5:  # 右辺
                cell.border = Border(top=no_border, left=no_border, right=thick, bottom=no_border)
            else:  # 内部 - 罫線なし
                cell.border = Border(top=no_border, left=no_border, right=no_border, bottom=no_border)
    
    # 下部の罫線
    for col in range(1, 6):
        cell = worksheet.cell(row=end_row, column=col)
        if col == 1:  # 左下角
            cell.border = Border(top=no_border, left=thick, right=no_border, bottom=thick)
        elif col == 5:  # 右下角
            cell.border = Border(top=no_border, left=no_border, right=thick, bottom=thick)
        else:  # 下辺
            cell.border = Border(top=no_border, left=no_border, right=no_border, bottom=thick)

if __name__ == "__main__":
    main()